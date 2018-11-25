#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on SAT Nov 4 2018

@author: Sara Palacios
"""

import openpyxl
from openpyxl import Workbook


class Excel:
    """
    Esta clase contiene el método inicializador, el que modifica el método str,
    el que lee un archivo y el que escribe en él
    """

    def __init__(self, file, sh, wb = False):

        """Este es el método inicializador que es llamado por el constructor.
            Si el tercer argumento es dado como 'True' se creará un nuevo
            archivo de excel,
            en caso de que aquel tercer argumento no sea dado o sea 'False',
            se dará por entendido que el archivo con el nombrre dado ya existe
            y por ende, simplemente se cargará.
        """
        
        self.filename = file
        if wb == True:
            self.wb = Workbook()#crea un documento nuevo
            self.sh = self.wb.active
            self.rws = self.sh.max_row #busca la última fila del documento
            self.clmns = self.sh.max_column#busca la última columna del documento
        else:
            self.wb = openpyxl.load_workbook(self.filename)#carga un documento ya existente
            self.sh = self.wb.active
            self.rws = self.sh.max_row
            self.clmns = self.sh.max_column

            
    def __str__(self):
        """Éste método modifica el método str original"""
        return "Usted está trabajando en el archivo: " + self.filename
    

    def read_wb(self, words1):
        """ Éste método lee el archivo y si la palabra dada se encuentra en él,
            entonces se crea una lista con el titulo del encabezado y
            su valor en la fila correspondiente.
        """
        
        info1=[]
               
                
        for i in range(1, self.rws+1):
            for j in range(1, self.clmns+1):
                a = self.sh.cell(i, j).value
                
                if a == words1:
                    for k in range(1, 20):
                        info1.append((self.sh.cell(1, k).value + "= "
                                     + str(self.sh.cell(i, k).value)))
        return info1
                
    def read_numbs(self, words1):
        """ Éste método lee el archivo y si la palabra dada se encuentra en él,
            entonces se crea una lista con los contenidos de la fila en la cual
            se encuentra esa palabra
        """
        
        info2=[]
               
                
        for i in range(1, self.rws+1):
            for j in range(1, self.clmns+1):
                a = self.sh.cell(i, j).value
                
                if a == words1:
                    for k in range(1, 20):
                        info2.append(self.sh.cell(i, k).value)
        return info2
        
        
               
                    
        
                

    def write_wb(self, lista):
        """ En éste método el usuario ingresa la información que desea insertar en
            en una fila del documento; ésta información es guardada en una lista y
            luego es escrita en el archivo
        """
        
        
                
        for j in range(self.rws+1, self.rws+2):            
            for k in range(1, len(lista)+1):
                self.sh.cell(j, k).value = lista[k-1]                
        self.wb.save(self.filename)#se guarda el archivo con la nueva información









